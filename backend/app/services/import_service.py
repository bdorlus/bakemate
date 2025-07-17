import uuid
from typing import Dict, Any
from pathlib import Path
from datetime import datetime

from sqlmodel import Session
from openpyxl import load_workbook

from app.models import (
    ImportJob,
    ImportState,
    IngredientCreate,
    ExpenseCreate,
    SupplyCreate,
    ExpenseCategory,
)
from app.services.ingredient_service import IngredientService
from app.services.recipe_service import RecipeService
from app.services.expense_service import ExpenseService
from app.services.supply_service import SupplyService
from app.models.user import User


class ImportService:
    def __init__(self, session: Session):
        self.session = session
        self.ingredient_service = IngredientService(session=session)
        self.recipe_service = RecipeService(session=session)
        self.expense_service = ExpenseService(session=session)
        self.supply_service = SupplyService(session=session)
        # Contacts and Orders not fully implemented in services; skipped

    async def create_job(self, user_id: uuid.UUID, file_path: Path) -> ImportJob:
        job = ImportJob(user_id=user_id, state=ImportState.QUEUED)
        self.session.add(job)
        self.session.commit()
        self.session.refresh(job)
        return job

    async def process_job(self, job_id: uuid.UUID, file_path: Path) -> None:
        job = self.session.get(ImportJob, job_id)
        if not job:
            return
        user = self.session.get(User, job.user_id)
        if not user:
            return
        job.state = ImportState.PROCESSING
        job.started_at = datetime.utcnow()
        job.summary = {}
        self.session.add(job)
        self.session.commit()

        summary: Dict[str, Dict[str, int]] = {
            "ingredients": {"created": 0, "errors": 0},
            "expenses": {"created": 0, "errors": 0},
            "supplies": {"created": 0, "errors": 0},
        }
        try:
            wb = load_workbook(file_path)
            if "Ingredients" in wb.sheetnames:
                sheet = wb["Ingredients"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    name, unit, price = row[:3]
                    ingredient_in = IngredientCreate(
                        user_id=job.user_id,
                        name=str(name).strip(),
                        unit=str(unit).strip(),
                        unit_cost=float(price or 0),
                    )
                    await self.ingredient_service.create_ingredient(
                        ingredient_in=ingredient_in, current_user=user
                    )
                    summary["ingredients"]["created"] += 1
            if "Expenses" in wb.sheetnames:
                sheet = wb["Expenses"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    date_val, category, amount, vendor, desc = row[:5]
                    exp_category = None
                    if category:
                        try:
                            exp_category = ExpenseCategory(category.lower())
                        except ValueError:
                            exp_category = ExpenseCategory.OTHER
                    expense_in = ExpenseCreate(
                        user_id=job.user_id,
                        date=date_val,
                        description=str(desc or ""),
                        amount=float(amount or 0),
                        category=exp_category,
                        vendor=vendor,
                    )
                    await self.expense_service.create_expense(
                        expense_in=expense_in, current_user=user, receipt_file=None
                    )
                    summary["expenses"]["created"] += 1
            if "Supplies" in wb.sheetnames:
                sheet = wb["Supplies"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    name, category, cost, qty = row[:4]
                    supply_in = SupplyCreate(
                        user_id=job.user_id,
                        name=str(name),
                        category=str(category) if category else None,
                        unit_cost=float(cost or 0),
                        stock_quantity=float(qty or 0),
                    )
                    await self.supply_service.create_supply(
                        supply_in=supply_in, current_user=user
                    )
                    summary["supplies"]["created"] += 1
            job.state = ImportState.COMPLETED
        except Exception:
            job.state = ImportState.FAILED
        finally:
            job.ended_at = datetime.utcnow()
            job.summary = summary
            self.session.add(job)
            self.session.commit()
