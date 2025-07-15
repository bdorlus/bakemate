import uuid
import time
from pathlib import Path
from fastapi.testclient import TestClient
from openpyxl import Workbook
from main import app

client = TestClient(app)


def create_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingredients"
    ws.append(["IngredientName", "MeasurementShort", "Price"])
    ws.append(["Sugar", "g", 1.2])
    ws.append(["Flour", "cup", 0.8])
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["ExpenseDate", "Category", "Amount", "Vendor", "Description"])
    ws2.append(["2024-01-01", "Supplies", 50, "Vendor1", "Sugar purchase"])
    ws2.append(["2024-01-02", "Utilities", 20, "Vendor2", "Electric bill"])
    ws3 = wb.create_sheet("Supplies")
    ws3.append(["Name", "Category", "Cost", "Quantity"])
    ws3.append(["Box", "Packaging", 0.1, 100])
    ws3.append(["Ribbon", "Packaging", 0.05, 50])
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


def test_full_import_flow(tmp_path):
    email = f"user{uuid.uuid4()}@example.com"
    password = "pass123"
    res = client.post(
        "/api/v1/auth/register", json={"email": email, "password": password}
    )
    assert res.status_code == 201

    res = client.post(
        "/api/v1/auth/login/access-token",
        data={"username": email, "password": password},
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    assert res.status_code == 200
    token = res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    workbook = create_workbook(tmp_path)
    with open(workbook, "rb") as f:
        res = client.post(
            "/api/v1/imports/",
            files={
                "file": (
                    "sample.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=headers,
        )
    assert res.status_code == 201
    job_id = res.json()["job_id"]

    for _ in range(5):
        status_res = client.get(f"/api/v1/imports/{job_id}", headers=headers)
        assert status_res.status_code == 200
        if status_res.json()["status"] == "completed":
            break
        time.sleep(0.1)
    data = status_res.json()
    assert data["status"] == "completed"
    assert data["summary"]["ingredients"]["created"] == 2
    assert data["summary"]["expenses"]["created"] == 2
    assert data["summary"]["supplies"]["created"] == 2
