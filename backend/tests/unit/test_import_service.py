from pathlib import Path
from sqlmodel import Session
from openpyxl import Workbook

from app.services.import_service import ImportService
from app.models import User, Ingredient, Expense, Supply
from sqlmodel import create_engine


engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})


def create_sample_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingredients"
    ws.append(["IngredientName", "MeasurementShort", "Price"])
    ws.append(["Sugar", "g", 1.2])
    ws.append(["Flour", "cup", 0.8])
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["ExpenseDate", "Category", "Amount", "Vendor", "Description"])
    ws2.append(["2024-01-01", "Supplies", 50, "Vendor1", "Sugar purchase"])
    ws2.append(["2024-01-02", "Utilities", 20, "Vendor2", "Electric bill"])
    ws3 = wb.create_sheet("Supplies")
    ws3.append(["Name", "Category", "Cost", "Quantity"])
    ws3.append(["Box", "Packaging", 0.1, 100])
    ws3.append(["Ribbon", "Packaging", 0.05, 50])
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


def test_import_creates_records(tmp_path):
    sample = create_sample_workbook(tmp_path)
    from sqlmodel import SQLModel

    SQLModel.metadata.drop_all(engine)
    SQLModel.metadata.create_all(engine)
    with Session(engine) as session:
        import uuid

        user = User(email=f"tester-{uuid.uuid4()}@example.com", hashed_password="x")
        session.add(user)
        session.commit()
        session.refresh(user)
        service = ImportService(session=session)
        import asyncio

        job = asyncio.run(service.create_job(user_id=user.id, file_path=sample))
        asyncio.run(service.process_job(job.id, sample))
        assert session.query(Ingredient).count() == 2
        assert session.query(Expense).count() == 2
        assert session.query(Supply).count() == 2
